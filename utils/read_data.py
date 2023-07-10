import openpyxl
def read_first_column(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    first_column = []
    for row in sheet.iter_rows(values_only=True):
        first_column.append(row[0])
    return first_column

def read_second_column(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    second_column = []
    for row in sheet.iter_rows(values_only=True):
        second_column.append(row[1])
    return second_column